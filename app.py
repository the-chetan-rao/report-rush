import platform
import shutil
from PyQt5.QtWidgets import QApplication, QMainWindow, QVBoxLayout, QMessageBox, QFileDialog, QDialog, QPushButton, QLabel
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtCore import QRectF, Qt, QTimer, QRect, QSize
from PyQt5.QtGui import QPainter, QColor, QPen, QIcon, QFont
from gui import Ui_MainWindow
import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import os
import sys
import time
import datetime
import random
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import re
import webbrowser

from project_ui import Project


class AboutDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("About This Project")
        self.setFixedSize(400, 200)  # Set a fixed size for the dialog
        # Keep the dialog on top
        # self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.setModal(True)

        # Set up the layout
        layout = QVBoxLayout(self)

        title_label = QLabel("About", self)
        # Create the labels with necessary information
        license_label = QLabel("Released under GNU GPL.", self)
        project_label = QLabel(
            "Project by Chetan Rao. Code by Abdul Mannan.", self)
        repo_label = QLabel(
            'Repo: <a href="https://github.com/your_repo">GitHub Link</a>', self)
        terms_label = QLabel(
            'By using this software, you agree to the <a href="https://github.com/the-chetan-rao/report-rush/blob/main/LICENSE.txt">Terms of Service</a>', self)

        # Enable clickable links in QLabel
        repo_label.setOpenExternalLinks(True)
        terms_label.setOpenExternalLinks(True)

        # Customize font and alignment
        title_label.setAlignment(Qt.AlignCenter)
        license_label.setAlignment(Qt.AlignLeft)
        project_label.setAlignment(Qt.AlignLeft)
        repo_label.setAlignment(Qt.AlignLeft)
        terms_label.setAlignment(Qt.AlignLeft)

        font = QFont()
        font.setPointSize(10)
        license_label.setFont(font)
        project_label.setFont(font)
        repo_label.setFont(font)
        terms_label.setFont(font)
        font.setPointSize(16)
        title_label.setFont(font)

        # Add labels to the layout
        layout.addWidget(title_label)
        layout.addWidget(license_label)
        layout.addWidget(project_label)
        layout.addWidget(repo_label)
        layout.addWidget(terms_label)


class LoginThread(QThread):
    login_successful = pyqtSignal()
    login_failed = pyqtSignal(str)

    def __init__(self, email, password, main_window):
        super().__init__()
        self.email = email
        self.password = password
        self.main_window = main_window

    def run(self):
        try:
            self.main_window.init_driver()
            self.main_window.driver.get("https://www.semrush.com/")
            time.sleep(random.uniform(5, 8))

            # Wait until the element is clickable (up to 180 seconds)
            element = WebDriverWait(self.main_window.driver, 180).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="srf-header"]/div/div/div[2]/a[1]/span[1]'))
            )
            element.click()
            time.sleep(random.uniform(2, 4))

            # Wait until the email field is present and send email
            email_element = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="email"]'))
            )
            email_element.send_keys(self.email)

            # Wait until the password field is present and send password
            password_element = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="password"]'))
            )
            password_element.send_keys(self.password)

            # Find and click the login button
            login_button = WebDriverWait(self.main_window.driver, 180).until(
                EC.element_to_be_clickable(
                    (By.XPATH, '//*[@id="loginForm"]/button[2]/span'))
            )
            login_button.click()

            time.sleep(random.uniform(2, 4))

            # After submitting the login form, wait for a specific element visible only after login
            profile_icon = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="srf-dropdown-menu-profile-menu"]/li[5]/a[3]'))
            )
            self.login_successful.emit()

        except Exception as e:
            self.login_failed.emit(str(e))
            if self.main_window.driver:
                self.main_window.driver.quit()


class ProcessThread(QThread):
    finished = pyqtSignal()
    error = pyqtSignal(str)
    progress = pyqtSignal(int)

    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window

    def scroll_until_all_errors_loaded(self):

        # Define the initial length of the errors list and set a scrolling loop
        prev_len = 0
        while True:
            # Scroll down to the bottom of the page
            self.main_window.driver.execute_script("""
        let scrollHeight = document.body.scrollHeight;
        let scrollPosition = 0;
        let scrollStep = 100; // Adjust this value for the step size (in pixels)
        let scrollInterval = setInterval(() => {
            window.scrollBy(0, scrollStep);
            scrollPosition += scrollStep;
            if (scrollPosition >= scrollHeight || window.innerHeight + window.scrollY >= scrollHeight) {
                clearInterval(scrollInterval); // Stop scrolling when reaching the bottom
            }
        }, 100); // Adjust this value for scroll speed (in milliseconds)
    """)

            # Adjust sleep time if needed to allow for content to load
            time.sleep(random.uniform(3, 6))

            # Retrieve all error elements currently loaded
            # errors = self.driver.find_elements(
            #     By.CSS_SELECTOR, '.sa-page-island-content.js-issueItem')

            errors = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_all_elements_located(
                    (By.CSS_SELECTOR, '.sa-page-island-content.js-issueItem:not(.sa-zeroIssues)'))
            )

            # If no new elements have been loaded, break the loop
            if len(errors) == prev_len:
                break
            prev_len = len(errors)  # Update the length for the next iteration

        # Collect all error URLs after scrolling is complete
        error_links = [error.find_element(
            By.TAG_NAME, 'a').get_attribute('href') for error in errors]
        print(f"Found {len(error_links)} files to download.")

        return error_links

    def run(self):
        try:

            if not self.main_window.projects:
                self.error.emit("Please add at least one project to start.")
                return
            if not self.main_window.path:
                self.error.emit("Please select a path to save the reports.")
                return
            if not self.main_window.is_chrome_installed():
                self.error.emit(
                    "Google Chrome is not installed on this system.")
                return

            self.main_window.driver.get("https://www.semrush.com/")
            time.sleep(random.uniform(4, 6))

            # Initialize lists
            filtered_divs = []
            elements = []

            urls = self.main_window.read_urls_from_file()

            # Fetching all div elements
            divs = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH,
                     "//div[@data-ui-name='Flex' and @role='cell' and contains(@class, '___SFlex_n3dnq-seo-d-sky') and @headers='igc-table-ui-kit-r5-projectName']")
                )
            )

            # Fetching all setup_audit elements
            setup_audit = WebDriverWait(self.main_window.driver, 180).until(
                EC.presence_of_all_elements_located(
                    (By.XPATH, "//a[@data-ui-name='Button' and contains(@class, '___SButton_1v3rj-seo-d-sky') and @type='button' and contains(@aria-label, 'Site Audit tool')]"))
            )

            # Loop through each div and check if the specific path exists
            for index, div in enumerate(divs):

                children = div.find_elements(By.XPATH, ".//*")

                # Construct the XPath using the current index
                specific_xpath = f"//*[@id='igc-ui-kit-rl-scroll-container']/div/div/div[{index + 1}]/div[3]/div/a"

                # Check if the specific element exists within the current div
                audit_buttons = div.find_elements(By.XPATH, specific_xpath)

                element = children[4].get_attribute('href').strip()

                # If the element exists, skip adding this div to the filtered_divs
                if not audit_buttons:  # If no elements were found, keep this div
                    filtered_divs.append(div)

                    if element in urls:

                        elements.append(element)

                if audit_buttons:

                    if element in urls:

                        self.main_window.un_audited.append(element)

                audit_buttons = None

            # Now filtered_divs contains only those divs that do not have the specific element

            self.main_window.loading.setText("Getting Projects")

            self.main_window.hrefs = elements

            if not self.main_window.hrefs:
                self.error.emit(
                    "Added project/s not found on Semrush or not setup for audit.")
                return

            total_hrefs = len(self.main_window.hrefs)
            for index, href in enumerate(self.main_window.hrefs):
                self.main_window.driver.get(href)
                time.sleep(random.uniform(4, 6))

                id = href.split('/')[-2]
                self.main_window.loading.setText(
                    f"Project id:{id}")

                site = f'https://www.semrush.com/siteaudit/campaign/{id}/review/overview/'

                self.main_window.driver.get(site)
                time.sleep(random.uniform(4, 6))

                button_issues = WebDriverWait(self.main_window.driver, 180).until(
                    EC.visibility_of_element_located(
                        (By.XPATH, '//*[@id="siteaudit-root"]/section/nav/div/button[2]'))
                )
                button_issues.click()

                time.sleep(random.uniform(7, 10))

                errors = self.scroll_until_all_errors_loaded()
                self.main_window.loading.setText(
                    f"Getting issues")

                for error in errors:
                    self.main_window.driver.get(error)
                    time.sleep(random.uniform(5, 7))

                    button_export = WebDriverWait(self.main_window.driver, 180).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="igc-ui-kit-ri-trigger"]'))
                    )
                    button_export.click()
                    time.sleep(random.uniform(4, 5))

                    button_download = WebDriverWait(self.main_window.driver, 180).until(
                        EC.element_to_be_clickable(
                            (By.XPATH, '//*[@id="igc-ui-kit-ri-popper"]/div/div/div/div[3]/button[1]'))
                    )
                    button_download.click()
                    self.main_window.loading.setText(
                        f"Downloading report")

                    self.main_window.move_latest_downloaded_file(
                        self.main_window.path, id)

                self.main_window.consolidate_excel_files(id)

                # Update progress
                self.progress.emit(int((index + 1) / total_hrefs * 100))

            self.finished.emit()

        except Exception as e:
            self.error.emit(str(e))


class LoadingIndicator(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)
        # Increased size for larger loading indicator
        self.setFixedSize(250, 250)
        self.setWindowFlags(Qt.FramelessWindowHint | Qt.Dialog)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.angle = 0
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.rotate)
        self.timer.start(50)
        self.text = "Logging in"  # Default loading text

    def rotate(self):
        self.angle += 10
        if self.angle >= 360:
            self.angle = 0
        self.update()

    def setText(self, text):
        """Set the loading text dynamically."""
        self.text = text
        self.update()  # Update the display to show the new text

    # def paintEvent(self, event):
    #     painter = QPainter(self)
    #     painter.setRenderHint(QPainter.Antialiasing)

    #     # Translate to the center for rotation
    #     painter.translate(self.width() / 2, self.height() / 2)
    #     painter.rotate(self.angle)

    #     # Draw rotating arc with increased size
    #     painter.setPen(QPen(QColor(0, 123, 255), 6,
    #                    Qt.SolidLine))  # Thicker arc
    #     painter.drawArc(QRectF(-40, -40, 80, 80), 0,
    #                     300 * 16)  # Increased arc size

    #     # Reset transformation for centered text drawing
    #     painter.resetTransform()

    #     # Set up font and color for the text
    #     font = painter.font()
    #     font.setPointSize(7)  # Larger font size for the text
    #     painter.setFont(font)
    #     text_rect = self.rect()

    #     # Measure the text to get its dimensions
    #     text_width = painter.fontMetrics().width(self.text)
    #     text_height = painter.fontMetrics().height()

    #     # Define the rectangle for the white background
    #     background_rect = QRect(
    #         0,   # X position with padding
    #         (self.height() - text_height) // 2 - 5,  # Y position with padding
    #         self.width(),                         # Width with padding
    #         text_height + 10                         # Height with padding
    #     )

    #     # Draw white background for the text
    #     painter.setBrush(QColor(255, 255, 255))  # White background
    #     painter.setPen(Qt.NoPen)                 # No border for the background
    #     painter.drawRect(background_rect)

    #     # Draw the text on top of the white background
    #     painter.setPen(QColor(0, 0, 0))          # Set text color to black
    #     painter.drawText(text_rect, Qt.AlignCenter, self.text)

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        # Translate to the center for rotation
        painter.translate(self.width() / 2, self.height() / 2)
        painter.rotate(self.angle)

        # Draw rotating arc with increased size
        painter.setPen(QPen(QColor(0, 123, 255), 6,
                       Qt.SolidLine))  # Thicker arc
        painter.drawArc(QRectF(-40, -40, 80, 80), 0,
                        300 * 16)  # Increased arc size

        # Reset transformation for centered text drawing
        painter.resetTransform()

        # Set up font and color for the text
        font = painter.font()
        font.setPointSize(7)  # Larger font size for the text
        painter.setFont(font)

        # Calculate dimensions for the text and background centered within arc boundary
        text_width = painter.fontMetrics().width(self.text)
        text_height = painter.fontMetrics().height()

        # Define the rectangle for the white background within arc boundary
        arc_boundary = 80  # Match the diameter of the arc for positioning
        background_rect = QRect(
            (self.width() - arc_boundary) // 2 -
            3,            # X start within arc
            (self.height() - text_height) // 2 - 5,        # Center vertically
            arc_boundary + 6,                                  # Width matching arc
            text_height + 10                               # Height with padding
        )

        # Draw white background for the text within arc boundary
        painter.setBrush(QColor(255, 255, 255))  # White background
        painter.setPen(Qt.NoPen)                 # No border for the background
        painter.drawRoundedRect(background_rect, 5, 5)

        # Draw the text on top of the white background
        painter.setPen(QColor(0, 0, 0))          # Set text color to black
        painter.drawText(background_rect, Qt.AlignCenter, self.text)

    def stop(self):
        self.timer.stop()
        self.close()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self):
        super().__init__()
        self.setupUi(self)  # Set up the UI from the .ui file

        # Initialize the scroll layout and path variable
        self.scroll_layout = None
        self.path = None
        self.projects = []
        self.email = None
        self.sanitized_email_as_file_name = None
        self.password = None
        self.driver = None
        self.hrefs = []
        self.process_thread = None
        self.loading = None
        self.un_audited = []

        # Set the window title and geometry
        self.setGeometry(150, 50, 600, 600)

        # Connect the add button to the add_project method
        self.ButtonAdd.clicked.connect(self.add_project)

        # Set the minimum height of the scroll area and set the layout
        self.scrollArea.setMinimumHeight(200)
        self.setLayoutForScroll()

        # Connect the path button to the set_path method
        self.ButtonPath.clicked.connect(self.set_path)

        # Connect the start button to the startProcess method
        self.ButtonStart.clicked.connect(self.startProcess)

        # Load the path from the txt file
        self.load_path_from_file()

        # Connect the login button to the login method
        self.ButtonLogin.clicked.connect(self.login)

        # Connect the textChanged signal to the custom function
        self.lineEditLoc.editingFinished.connect(self.on_text_changed_set_path)

        self.stackedWidget.setCurrentIndex(1)

        # Add the about button to the main window
        self.add_about_button()

        self.labelMinimize.hide()

        # Open X
        self.ButtonX.clicked.connect(self.openX)

        self.about_button.clicked.connect(self.show_about_dialog)

    def sanitize_email_for_filename(self, email):
        # Replace '@' with '_at_' and '.' with '_dot_'
        sanitized_email = email.replace('@', '_at_').replace('.', '_dot_')
        return sanitized_email

    def show_about_dialog(self):
        # Show the About dialog centered on the main window
        about_dialog = AboutDialog(self)
        about_dialog.exec_()

    def openX(self):
        self.ButtonX.setCursor(Qt.PointingHandCursor)
        # Open the specified URL in the default browser
        webbrowser.open("https://x.com/thechetanrao")

    def update_about_button_position(self):
        # Calculate the position of the about button
        title_bar_height = 5
        right_margin = 10

        button_x = self.centralWidget().width() - self.about_button.width() - right_margin
        button_y = title_bar_height + 10
        self.about_button.setGeometry(QRect(button_x, button_y, 24, 24))

    def add_about_button(self):

        self.about_button = QPushButton(self)
        self.about_button.setIcon(QIcon(":/icons/question-mark.png"))
        self.about_button.setIconSize(QSize(24, 24))
        self.about_button.setToolTip("About")
        self.about_button.setFlat(True)

        # Set the button stylesheet
        self.about_button.setStyleSheet("""
            QPushButton {
                background-color: transparent; /* Default background */
                border: none; /* No border */
                padding: 5px; /* Add padding for better clickable area */
            }
            QPushButton:hover {
                border: 2px solid rgba(255, 255, 255, 0.8); /* Light border effect on hover */
            }
            QPushButton:pressed {
                border: 2px solid rgba(0, 0, 0, 0.5); /* Darker border effect on pressed */
                border-radius: 12px; /* Round the button corners */
                background-color: rgba(255, 255, 255, 0.2); /* Slightly darken background when pressed */
            }
        """)

        self.update_about_button_position()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        # Update the loading indicator position whenever the window is resized
        self.center_loading_indicator()
        self.update_about_button_position()

    def center_loading_indicator(self):
        if self.loading:
            # Calculate the center position of the main window
            center = self.centralWidget().rect().center()

            # Move the loading indicator to the center
            global_center = self.centralWidget().mapToGlobal(center)
            self.loading.move(self.mapFromGlobal(
                global_center) - self.loading.rect().center())

    def login(self):

        self.email = self.lineEditEmail.text()
        self.password = self.lineEditPassword.text()

        if self.email == "" or self.password == "":
            QMessageBox.warning(self, "No Credentials",
                                "Please enter your email and password.")
            return

        self.loading = LoadingIndicator()
        self.loading.setParent(self)
        self.loading.show()
        self.loading.adjustSize()
        self.center_loading_indicator()
        # Change the loading text as needed
        self.loading.setText("Logging in")

        # Set focus to the loading indicator
        self.loading.setFocus()

        self.login_thread = LoginThread(self.email, self.password, self)
        self.login_thread.login_successful.connect(self.onLoginSuccess)
        self.login_thread.login_failed.connect(self.onLoginFailed)
        self.login_thread.start()

        # Disable login button
        self.ButtonLogin.setEnabled(False)

    def onLoginSuccess(self):
        self.loading.stop()
        self.loading = None
        QMessageBox.information(self, "Login Successful",
                                "You have successfully logged in.")
        self.stackedWidget.setCurrentIndex(0)
        # Enable login button or hide loading indicator
        self.ButtonLogin.setEnabled(True)

        # Set the sanitized email as the file name
        self.sanitized_email_as_file_name = self.sanitize_email_for_filename(
            self.email)
        self.sanitized_email_as_file_name += ".txt"
        # Load existing projects from the text file
        self.load_projects_from_file()

    def onLoginFailed(self, error_message):
        # QMessageBox.warning(self, "Login Failed",
        #                     f"Please check your email and password.")
        QMessageBox.warning(self, "Login Failed", f"{error_message}")
        # Enable login button or hide loading indicator
        self.ButtonLogin.setEnabled(True)
        self.loading.stop()
        self.loading = None

    def startProcess(self):

        self.loading = LoadingIndicator()
        self.loading.setParent(self)
        self.loading.show()
        self.loading.adjustSize()
        self.center_loading_indicator()
        # Change the loading text as needed
        self.loading.setText("Processing")

        # Set focus to the loading indicator
        self.loading.setFocus()

        self.process_thread = ProcessThread(self)
        self.process_thread.finished.connect(self.onProcessFinished)
        self.process_thread.error.connect(self.onProcessError)
        self.process_thread.progress.connect(self.updateProgress)
        self.process_thread.start()

        self.ButtonStart.setEnabled(False)

        self.labelMinimize.show()

    def onProcessFinished(self):

        # Remove duplicates
        self.un_audited = list(set(self.un_audited))
        self.labelMinimize.hide()

        str = 'Projects with id: '
        for href in self.un_audited:
            id = href.split('/')[-2]
            str = str + f'{id},'
        str = str[:-2]
        QMessageBox.information(
            self, "Projects Skipped", f"{str} skipped due to not being audited.")

        self.ButtonStart.setEnabled(True)
        QMessageBox.information(self, "Process Completed",
                                "The process has finished successfully.")

        self.loading.stop()
        self.loading = None

    def onProcessError(self, error_message):

        self.labelMinimize.hide()

        self.ButtonStart.setEnabled(True)
        QMessageBox.warning(self, "Error", error_message)

        self.loading.stop()
        self.loading = None

    def updateProgress(self, value):
        # Update progress bar or status message here

        print(f"Progress: {value}%")

    def load_path_from_file(self):
        """Load the path from the text file and set it in the line edit."""
        try:
            with open('path.txt', 'r') as file:
                path = file.read().strip()
            self.lineEditLoc.setText(path)
            self.path = path
        except FileNotFoundError:
            pass

    def init_driver(self):

        # Create a ChromeOptions object
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_argument("--blink-settings=imagesEnabled=false")
        chrome_options.add_argument("--window-position=-10000,0")
        chrome_options.add_argument('--disable-gpu')
        chrome_options.add_argument("--start-minimized")
        chrome_options.add_argument('--enable-logging')
        chrome_options.add_argument('--v=1')
        chrome_options.add_argument(
            "--safebrowsing-disable-download-protection")
        chrome_options.add_argument(
            "--safebrowsing-disable-extension-blacklist")

        # Ensure ChromeDriver is compatible with the installed Chrome version
        self.driver = uc.Chrome(
            driver_executable_path=ChromeDriverManager().install(), options=chrome_options)

        # Return the driver
        return self.driver

    def on_text_changed_set_path(self):
        text = self.lineEditLoc.text().strip()
        """Set the path when the text in the line edit changes."""
        if os.path.exists(text):
            self.path = text

            # Save path to txt file
            with open('path.txt', 'w') as file:
                file.write(text)

        else:
            self.path = None
            QMessageBox.warning(self, "Invalid Path",
                                "The specified path does not exist.")

    def load_projects_from_file(self):
        """Load existing projects from the text file and create project widgets."""
        urls = self.read_urls_from_file()
        for url in urls:
            # Use the add_project method to create and add each project
            self.add_project(url)

    def setLayoutForScroll(self):
        # Set layout for scrollArea if not set yet
        if self.scroll_layout is None:
            self.scroll_layout = QVBoxLayout(self.scrollAreaWidgetContents)
            self.scroll_layout.addStretch()
            self.scroll_layout.setContentsMargins(0, 0, 9, 0)

    def set_path(self):
        """Sets the folder path for the output file"""
        options = QFileDialog.Options()

        # Open a dialog to select a folder
        folder = QFileDialog.getExistingDirectory(
            self, "Select Folder", "", options=options)

        # If a folder is selected, set it in the line edit
        if folder:
            self.lineEditLoc.setText(folder)
            self.path = folder

            # Save path to txt file
            with open('path.txt', 'w') as file:
                file.write(folder)

    def remove_project(self, project):
        """Removes a project widget from the layout and the list"""
        self.scroll_layout.removeWidget(project)  # Remove from layout
        project.deleteLater()  # Ensure widget is deleted properly
        self.projects.remove(project)  # Remove from list

        # Remove the URL from the file
        url = project.url  # Make sure the Project class has a way to access its URL
        self.remove_url_from_file(url)

    def add_project(self, url=None):

        if url == False:
            url = self.lineEditURL.text()

        if url.strip() == "":
            # Show a message box if URL is empty
            msg_box = QMessageBox(self)
            msg_box.setIcon(QMessageBox.Warning)
            msg_box.setWindowTitle("Input Error")
            msg_box.setText("Please enter a URL")
            msg_box.setStandardButtons(QMessageBox.Ok)
            msg_box.exec_()
            return  # Stop execution if the input is empty

        url = url.strip()  # Remove leading and trailing whitespace
        # Create a new Project widget
        project = Project(url)

        # Connect the remove button to the remove_project method
        project.removeButton.clicked.connect(
            lambda: self.remove_project(project))

        # Insert the project widget before the stretch (at index -1)
        self.scroll_layout.insertWidget(
            self.scroll_layout.count() - 1, project)

        # Add the project to the list
        self.projects.append(project)

        # Write the URL to the file if added from the line edit
        if url == self.lineEditURL.text().strip():
            self.lineEditURL.clear()
            self.write_url_to_file(url)

    def is_chrome_installed(self):
        """Check if Google Chrome is installed based on the operating system."""
        if os.name == "nt":  # Windows
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
            ]
            return any(os.path.exists(path) for path in chrome_paths)

        elif sys.platform == "darwin":  # macOS
            return os.path.exists("/Applications/Google Chrome.app")

        elif sys.platform.startswith("linux"):  # Linux
            chrome_paths = [
                "/usr/bin/google-chrome",
                "/opt/google/chrome/chrome"
            ]
            return any(os.path.exists(path) for path in chrome_paths)

        else:
            raise RuntimeError("Unsupported operating system")

    def write_url_to_file(self, url):
        """Append the project URL to the text file."""
        with open(self.sanitized_email_as_file_name, 'a') as file:
            file.write(url + '\n')

    def remove_url_from_file(self, url):
        """Remove the project URL from the text file."""
        try:
            # Read all lines from the file
            with open(self.sanitized_email_as_file_name, 'r') as file:
                lines = file.readlines()

            # Write back all lines except the one that needs to be removed
            with open(self.sanitized_email_as_file_name, 'w') as file:
                for line in lines:
                    if line.strip() != url:
                        file.write(line)
        except FileNotFoundError:
            pass  # If the file does not exist, simply pass

    def read_urls_from_file(self):
        """Read project URLs from the text file and return them as a list."""
        try:
            with open(self.sanitized_email_as_file_name, 'r') as file:
                urls = [line.strip()
                        for line in file.readlines() if line.strip()]
            return urls
        except FileNotFoundError:
            return []  # Return an empty list if the file does not exist

    def get_default_download_folder(self):
        """Returns the default download folder path based on the operating system."""
        if platform.system() == 'Windows':
            return os.path.join(os.environ['USERPROFILE'], 'Downloads')
        elif platform.system() == 'Darwin':  # macOS
            return os.path.join(os.path.expanduser('~'), 'Downloads')
        else:  # Assume Linux or other UNIX-based systems
            return os.path.join(os.path.expanduser('~'), 'Downloads')

    def move_latest_downloaded_file(self, destination_folder, href):
        """Moves the latest downloaded file to the specified destination folder."""
        download_folder = self.get_default_download_folder()

        # Wait for a new file to appear and be fully downloaded
        latest_file = None
        while True:
            latest_file = self.get_latest_file(download_folder)
            # Make sure it's not a partially downloaded file
            if latest_file and not latest_file.endswith('.crdownload'):
                break
            print("Waiting for file to be fully downloaded...")
            time.sleep(1)

        # Move the latest file to the destination folder
        try:
            source_path = os.path.join(
                download_folder, latest_file)
            os.makedirs(os.path.join(destination_folder, href), exist_ok=True)
            destination_path = os.path.join(
                destination_folder, href, latest_file)

            shutil.move(source_path, destination_path)
            print(f"File {latest_file} moved to {destination_path}")
        except Exception as e:
            print(f"Error moving file: {e}")

    def get_latest_file(self, download_folder):
        """Returns the most recently modified file in the download folder within the last minute."""
        # Get the current time
        current_time = time.time()

        # List all files in the directory and check if they are files
        files = [f for f in os.listdir(download_folder) if os.path.isfile(
            os.path.join(download_folder, f))]

        if not files:
            return None

        # Filter files modified in the last minute
        files_in_last_minute = [f for f in files if (
            current_time - os.path.getmtime(os.path.join(download_folder, f))) <= 60]

        if not files_in_last_minute:
            return None

        # Get the most recently modified file in the last minute
        latest_file = max(files_in_last_minute, key=lambda f: os.path.getmtime(
            os.path.join(download_folder, f)))

        return latest_file

    def find_common_prefix(self, filenames):
        if not filenames:
            return ""
        s1 = min(filenames)
        s2 = max(filenames)
        for i, c in enumerate(s1):
            if c != s2[i]:
                return s1[:i]
        return s1

    def process_sheet_name(self, name, common_prefix):
        # Remove common prefix, date string, and .xlsx extension
        name = name.removeprefix(common_prefix)
        name = re.sub(r'_\d{8}\.xlsx$', '', name)

        # Remove underscores and convert to camel case
        name = ''.join(word.capitalize() for word in name.split('_'))
        name = name[0].lower() + name[1:]

        # Truncate to 28 characters if necessary
        return name[:28]

    def make_unique_sheet_name(self, workbook, sheet_name):
        # Truncate to 28 characters to leave room for suffix
        base_name = sheet_name[:28]
        if base_name not in workbook.sheetnames:
            return base_name

        suffix = 1
        while True:
            # Ensure total length doesn't exceed 31
            new_name = f"{base_name[:27]}{suffix}"
            if new_name not in workbook.sheetnames:
                return new_name
            suffix += 1

    def consolidate_excel_files(self, id):
        try:
            current_dir = os.path.join(self.path, id)
            excel_files = [f for f in os.listdir(current_dir) if f.endswith(
                '.xlsx') and f != 'consolidated_output.xlsx']

            if not excel_files:
                print("No Excel files found in the current directory.")
                return

            print(
                f"Found {len(excel_files)} Excel files. Starting consolidation process...")

            # Find common prefix
            common_prefix = self.find_common_prefix(excel_files)
            print(f"Common prefix found: '{common_prefix}'")

            consolidated_workbook = openpyxl.Workbook()
            contents_sheet = consolidated_workbook.active
            contents_sheet.title = "Contents"
            contents_sheet['A1'] = "Original Filename"
            contents_sheet['B1'] = "Sheet Name"
            contents_sheet['A1'].font = Font(bold=True)
            contents_sheet['B1'].font = Font(bold=True)

            self.loading.setText(
                f"Consolidating Files")

            row = 2
            for filename in excel_files:
                print(f"Processing file: {filename}")
                sheet_name = self.process_sheet_name(filename, common_prefix)
                sheet_name = self.make_unique_sheet_name(
                    consolidated_workbook, sheet_name)

                # Read the Excel file
                df = pd.read_excel(os.path.join(current_dir, filename))

                # Remove the "Discovered" column if it exists
                if 'Discovered' in df.columns:
                    df = df.drop('Discovered', axis=1)

                # Create a new sheet and write data
                sheet = consolidated_workbook.create_sheet(sheet_name)

                # Write headers starting from column B
                for c, header in enumerate(df.columns, start=2):
                    sheet.cell(row=1, column=c, value=header)
                    sheet.cell(row=1, column=c).font = Font(bold=True)

                # Write data starting from column B
                for r, row_data in enumerate(df.values, start=2):
                    for c, value in enumerate(row_data, start=2):
                        sheet.cell(row=r, column=c, value=value)

                # Add link back to contents
                sheet['A1'] = 'Back to Contents'
                sheet['A1'].hyperlink = f"#Contents!A1"
                sheet['A1'].font = Font(color="0000FF", underline="single")

                # Add entry to contents sheet
                contents_sheet.cell(row=row, column=1, value=filename)
                contents_sheet.cell(row=row, column=2, value=sheet_name)
                contents_sheet.cell(
                    row=row, column=2).hyperlink = f"#{sheet_name}!A1"
                contents_sheet.cell(row=row, column=2).font = Font(
                    color="0000FF", underline="single")
                row += 1

            # Adjust column widths in contents sheet
            for column in contents_sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                contents_sheet.column_dimensions[column_letter].width = adjusted_width

            date_time = datetime.datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
            output_path = f"consolidated_output_{date_time}.xlsx"
            output_path = os.path.join(current_dir, output_path)
            consolidated_workbook.save(output_path)
            print(
                f"Consolidation completed successfully. Output file: {output_path}")

        except Exception as e:
            print(f"An error occurred: {str(e)}")
            print("Please ensure you have the necessary permissions and that the files are not open in another program.")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
